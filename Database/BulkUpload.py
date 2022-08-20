import mysql.connector
from openpyxl import load_workbook, Workbook
import mysql.connector as connection

# wb = load_workbook("D:\My Videos\Programming Language\Python\Library\Hindi\Python(Openpyxl and MYSQL) Hindi Bulk data upload to database using python from excel\Insurance Data.xlsx")
#
# ws = wb['PolicyData']

conn = connection.connect(host='localhost', user='root', password="Deva@321#", database='webchamp')

cursor = conn.cursor()

#cursor.execute("""create table insurance(Policy int,Expiry Date,Location varchar(100),State varchar(100),Region varchar(100),InsuredValue int,Construction varchar(100),BusinessType varchar(100),Earthquake varchar(100),Flood varchar(100))""")

#insert into insurance values(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9])

# for row in ws.values:
    # if row[0]=="Policy":
    #     pass
    # else:
    #     cursor.execute(f"""insert into insurance values({row[0]},"{row[1]}","{row[2]}","{row[3]}","{row[4]}",{row[5]},"{row[6]}","{row[7]}","{row[8]}","{row[9]}")""")
    #     conn.commit()

    # if type(row[0])==int:
    #     cursor.execute(f"""insert into insurance values({row[0]},"{row[1]}","{row[2]}","{row[3]}","{row[4]}",{row[5]},"{row[6]}","{row[7]}","{row[8]}","{row[9]}")""")
    #     conn.commit()

    #print(type(row[0]))

cursor.execute("""select * from insurance""")

wb = Workbook()

ws = wb.active

raw_data = [("Policy", "Expiry", "Location", "State", "Region", "InsuredValue", "Construction", "BusinessType", "Earthquake", "Flood"),]

for i in cursor:
    raw_data.append(i)

for i in raw_data:
    ws.append(i)

wb.save("D:\My Videos\Programming Language\Python\Library\Hindi\Python(Openpyxl and MYSQL) Hindi Bulk data upload to database using python from excel\\test.xlsx")